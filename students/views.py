"""
Vues pour l'application students
"""
import re
import unicodedata
from datetime import date, datetime

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.utils import timezone
from .models import Student, Inscription, TypeDocument, DocumentEtudiant, DossierEtudiant
from .forms import StudentForm, StudentImportForm, InscriptionForm, TypeDocumentForm, DocumentEtudiantForm


# ========== ETUDIANTS ==========
@login_required
def student_list(request):
    students = Student.objects.all().order_by('numero_etudiant')
    paginator = Paginator(students, 15)
    page = request.GET.get('page')
    students = paginator.get_page(page)
    return render(request, 'students/student_list.html', {'students': students})


def _normalize_header(value):
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def _parse_excel_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        cleaned = value.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(cleaned, fmt).date()
            except ValueError:
                continue
    return None


def _normalize_sexe(value):
    text = str(value or "").strip().lower()
    if text in {"m", "masculin", "male", "homme"}:
        return "M"
    if text in {"f", "feminin", "féminin", "female", "femme"}:
        return "F"
    return ""


@login_required
def student_import(request):
    results = None
    errors = []

    if request.method == 'POST':
        form = StudentImportForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded = form.cleaned_data['fichier_excel']
            try:
                from openpyxl import load_workbook
            except ImportError:
                messages.error(request, "La bibliothèque openpyxl n'est pas installée sur le serveur.")
                return render(request, 'students/student_import.html', {'form': form, 'title': "Import Excel des étudiants"})

            try:
                workbook = load_workbook(uploaded, data_only=True)
            except Exception as exc:
                messages.error(request, f"Impossible de lire le fichier Excel : {exc}")
                return render(request, 'students/student_import.html', {'form': form, 'title': "Import Excel des étudiants"})

            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if len(rows) < 2:
                messages.warning(request, "Le fichier Excel ne contient aucune donnée exploitable.")
                return render(request, 'students/student_import.html', {'form': form, 'title': "Import Excel des étudiants"})

            headers = [_normalize_header(h) for h in rows[0]]
            index_map = {header: idx for idx, header in enumerate(headers) if header}

            required_headers = ['numero_etudiant', 'nom', 'prenom', 'date_naissance', 'lieu_naissance', 'sexe', 'email']
            missing_headers = [header for header in required_headers if header not in index_map]
            if missing_headers:
                messages.error(
                    request,
                    "Colonnes manquantes dans le fichier Excel : " + ", ".join(missing_headers)
                )
                return render(request, 'students/student_import.html', {'form': form, 'title': "Import Excel des étudiants"})

            created = 0
            updated = 0

            with transaction.atomic():
                for line_number, row in enumerate(rows[1:], start=2):
                    if not any(row):
                        continue

                    def cell(name):
                        idx = index_map.get(name)
                        return row[idx] if idx is not None and idx < len(row) else None

                    numero_etudiant = str(cell('numero_etudiant') or '').strip()
                    nom = str(cell('nom') or '').strip()
                    prenom = str(cell('prenom') or '').strip()
                    date_naissance = _parse_excel_date(cell('date_naissance'))
                    lieu_naissance = str(cell('lieu_naissance') or '').strip()
                    sexe = _normalize_sexe(cell('sexe'))
                    email = str(cell('email') or '').strip()

                    if not all([numero_etudiant, nom, prenom, date_naissance, lieu_naissance, sexe, email]):
                        errors.append(f"Ligne {line_number}: champs obligatoires manquants.")
                        continue

                    defaults = {
                        'nom': nom,
                        'prenom': prenom,
                        'date_naissance': date_naissance,
                        'lieu_naissance': lieu_naissance,
                        'sexe': sexe,
                        'email': email,
                    }

                    optional_values = {
                        'nationalite': str(cell('nationalite') or '').strip(),
                        'telephone': str(cell('telephone') or '').strip(),
                        'adresse': str(cell('adresse') or '').strip(),
                        'statut': str(cell('statut') or '').strip().lower(),
                    }

                    if optional_values['nationalite']:
                        defaults['nationalite'] = optional_values['nationalite']
                    if optional_values['telephone']:
                        defaults['telephone'] = optional_values['telephone']
                    if optional_values['adresse']:
                        defaults['adresse'] = optional_values['adresse']
                    if optional_values['statut']:
                        statut_value = optional_values['statut']
                        allowed_statuses = {'actif', 'suspendu', 'exclu', 'diplome', 'abandon'}
                        if statut_value in allowed_statuses:
                            defaults['statut'] = statut_value
                        else:
                            errors.append(
                                f"Ligne {line_number}: statut invalide '{optional_values['statut']}'."
                            )
                            continue

                    student, created_flag = Student.objects.update_or_create(
                        numero_etudiant=numero_etudiant,
                        defaults=defaults,
                    )
                    if created_flag:
                        created += 1
                    else:
                        updated += 1

            results = {'created': created, 'updated': updated, 'errors': errors}
            if created or updated:
                messages.success(
                    request,
                    f"Import terminé : {created} créé(s), {updated} mis à jour(s)."
                )
            if errors:
                messages.warning(request, f"{len(errors)} ligne(s) ont été ignorées.")

            return render(request, 'students/student_import.html', {
                'form': StudentImportForm(),
                'title': "Import Excel des étudiants",
                'results': results,
            })
    else:
        form = StudentImportForm()

    return render(request, 'students/student_import.html', {'form': form, 'title': 'Import Excel des étudiants'})


@login_required
def student_create(request):
    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Étudiant créé avec succès!')
            return redirect('students:student_list')
    else:
        form = StudentForm()
    return render(request, 'students/student_form.html', {'form': form, 'title': 'Nouvel Étudiant'})


@login_required
def student_update(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES, instance=student)
        if form.is_valid():
            form.save()
            messages.success(request, 'Étudiant modifié avec succès!')
            return redirect('students:student_list')
    else:
        form = StudentForm(instance=student)
    return render(request, 'students/student_form.html', {'form': form, 'title': 'Modifier Étudiant', 'object': student})


@login_required
def student_delete(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        student.delete()
        messages.success(request, 'Étudiant supprimé avec succès!')
        return redirect('students:student_list')
    return render(request, 'students/student_confirm_delete.html', {'student': student})


@login_required
def student_detail(request, pk):
    student = get_object_or_404(Student, pk=pk)
    inscriptions = student.inscriptions.all().order_by('-annee_academique')
    return render(request, 'students/student_detail.html', {'student': student, 'inscriptions': inscriptions})


# ========== INSCRIPTIONS ==========
@login_required
def inscription_list(request):
    inscriptions = Inscription.objects.select_related('etudiant', 'classe', 'classe__promotion', 'classe__promotion__filiere', 'classe__promotion__filiere__section', 'classe__local', 'annee_academique').all().order_by('-annee_academique', 'etudiant')
    paginator = Paginator(inscriptions, 15)
    page = request.GET.get('page')
    inscriptions = paginator.get_page(page)
    return render(request, 'students/inscription_list.html', {'inscriptions': inscriptions})


@login_required
def inscription_create(request):
    if request.method == 'POST':
        form = InscriptionForm(request.POST)
        if form.is_valid():
            inscription = form.save()
            # Créer automatiquement le dossier étudiant
            DossierEtudiant.objects.get_or_create(inscription=inscription)
            messages.success(request, 'Inscription créée avec succès!')
            return redirect('students:inscription_list')
    else:
        form = InscriptionForm()
    return render(request, 'students/inscription_form.html', {'form': form, 'title': 'Nouvelle Inscription'})


@login_required
def inscription_update(request, pk):
    inscription = get_object_or_404(Inscription, pk=pk)
    if request.method == 'POST':
        form = InscriptionForm(request.POST, instance=inscription)
        if form.is_valid():
            form.save()
            messages.success(request, 'Inscription modifiée avec succès!')
            return redirect('students:inscription_list')
    else:
        form = InscriptionForm(instance=inscription)
    return render(request, 'students/inscription_form.html', {'form': form, 'title': 'Modifier Inscription', 'object': inscription})


@login_required
def inscription_delete(request, pk):
    inscription = get_object_or_404(Inscription, pk=pk)
    if request.method == 'POST':
        inscription.delete()
        messages.success(request, 'Inscription supprimée avec succès!')
        return redirect('students:inscription_list')
    return render(request, 'students/inscription_confirm_delete.html', {'inscription': inscription})


# ========== DOCUMENTS ==========
@login_required
def document_list(request):
    documents = DocumentEtudiant.objects.select_related('etudiant', 'type_document').all().order_by('-date_depot')
    paginator = Paginator(documents, 15)
    page = request.GET.get('page')
    documents = paginator.get_page(page)
    return render(request, 'students/document_list.html', {'documents': documents})


@login_required
def document_create(request):
    if request.method == 'POST':
        form = DocumentEtudiantForm(request.POST, request.FILES)
        if form.is_valid():
            document = form.save(commit=False)
            if document.valide:
                document.date_validation = timezone.now().date()
                document.valide_par = request.user
            document.save()
            messages.success(request, 'Document ajouté avec succès!')
            return redirect('students:document_list')
    else:
        form = DocumentEtudiantForm()
    return render(request, 'students/document_form.html', {'form': form, 'title': 'Nouveau Document'})


@login_required
def document_update(request, pk):
    document = get_object_or_404(DocumentEtudiant, pk=pk)
    if request.method == 'POST':
        form = DocumentEtudiantForm(request.POST, request.FILES, instance=document)
        if form.is_valid():
            doc = form.save(commit=False)
            if doc.valide and not document.valide:
                doc.date_validation = timezone.now().date()
                doc.valide_par = request.user
            doc.save()
            messages.success(request, 'Document modifié avec succès!')
            return redirect('students:document_list')
    else:
        form = DocumentEtudiantForm(instance=document)
    return render(request, 'students/document_form.html', {'form': form, 'title': 'Modifier Document', 'object': document})


@login_required
def document_delete(request, pk):
    document = get_object_or_404(DocumentEtudiant, pk=pk)
    if request.method == 'POST':
        document.delete()
        messages.success(request, 'Document supprimé avec succès!')
        return redirect('students:document_list')
    return render(request, 'students/document_confirm_delete.html', {'document': document})


# ========== TYPES DE DOCUMENTS ==========
@login_required
def type_document_list(request):
    types = TypeDocument.objects.all().order_by('ordre', 'nom')
    return render(request, 'students/type_document_list.html', {'types': types})


@login_required
def type_document_create(request):
    if request.method == 'POST':
        form = TypeDocumentForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Type de document créé avec succès!')
            return redirect('students:type_document_list')
    else:
        form = TypeDocumentForm()
    return render(request, 'students/type_document_form.html', {'form': form, 'title': 'Nouveau Type de Document'})


@login_required
def type_document_update(request, pk):
    type_doc = get_object_or_404(TypeDocument, pk=pk)
    if request.method == 'POST':
        form = TypeDocumentForm(request.POST, instance=type_doc)
        if form.is_valid():
            form.save()
            messages.success(request, 'Type de document modifié avec succès!')
            return redirect('students:type_document_list')
    else:
        form = TypeDocumentForm(instance=type_doc)
    return render(request, 'students/type_document_form.html', {'form': form, 'title': 'Modifier Type de Document', 'object': type_doc})


@login_required
def type_document_delete(request, pk):
    type_doc = get_object_or_404(TypeDocument, pk=pk)
    if request.method == 'POST':
        type_doc.delete()
        messages.success(request, 'Type de document supprimé avec succès!')
        return redirect('students:type_document_list')
    return render(request, 'students/type_document_confirm_delete.html', {'type_doc': type_doc})


# ========== DOSSIERS ==========
@login_required
def dossier_list(request):
    dossiers = DossierEtudiant.objects.select_related('inscription__etudiant', 'inscription__classe', 'inscription__classe__promotion').all().order_by('-date_ouverture')
    paginator = Paginator(dossiers, 15)
    page = request.GET.get('page')
    dossiers = paginator.get_page(page)
    return render(request, 'students/dossier_list.html', {'dossiers': dossiers})


@login_required
def dossier_detail(request, pk):
    dossier = get_object_or_404(DossierEtudiant, pk=pk)
    documents = dossier.inscription.documents.all()
    return render(request, 'students/dossier_detail.html', {'dossier': dossier, 'documents': documents})
