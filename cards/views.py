from io import BytesIO
import re
import unicodedata

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.urls import reverse_lazy, reverse
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.db import transaction
from django.utils import timezone
from .models import Card, Personnel, Position, Category
from .forms import CardForm, PersonnelForm, PersonnelImportForm
from .utils import generate_card_image

# --- PERSONNEL VIEWS ---

def _normalize_header(value):
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def _normalize_text(value):
    return str(value or "").strip()


@login_required
@permission_required('cards.add_personnel', raise_exception=True)
def personnel_import(request):
    results = None

    if request.method == 'POST':
        form = PersonnelImportForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded = form.cleaned_data['fichier_excel']
            try:
                from openpyxl import load_workbook
            except ImportError:
                messages.error(request, "La bibliothèque openpyxl n'est pas installée sur le serveur.")
                return render(request, 'personnel/personnel_import.html', {'form': form, 'title': "Import Excel du personnel"})

            try:
                workbook = load_workbook(uploaded, data_only=True)
            except Exception as exc:
                messages.error(request, f"Impossible de lire le fichier Excel : {exc}")
                return render(request, 'personnel/personnel_import.html', {'form': form, 'title': "Import Excel du personnel"})

            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if len(rows) < 2:
                messages.warning(request, "Le fichier Excel ne contient aucune donnée exploitable.")
                return render(request, 'personnel/personnel_import.html', {'form': form, 'title': "Import Excel du personnel"})

            headers = [_normalize_header(h) for h in rows[0]]
            index_map = {header: idx for idx, header in enumerate(headers) if header}

            required_headers = ['prenom', 'nom', 'poste_fonction', 'categorie', 'niveau_d_etudes']
            alias_map = {
                'poste_fonction': ['poste_fonction', 'poste', 'fonction', 'position'],
                'categorie': ['categorie', 'cat', 'category'],
                'niveau_d_etudes': ['niveau_d_etudes', 'niveau_etudes', 'education_level'],
                'matricule': ['matricule', 'matricule_personnel'],
                'photo': ['photo', 'image', 'photo_de_profil'],
            }

            def get_index(name):
                for candidate in alias_map.get(name, [name]):
                    if candidate in index_map:
                        return index_map[candidate]
                return None

            missing_headers = [header for header in required_headers if get_index(header) is None]
            if missing_headers:
                messages.error(
                    request,
                    "Colonnes manquantes dans le fichier Excel : " + ", ".join(missing_headers)
                )
                return render(request, 'personnel/personnel_import.html', {'form': form, 'title': "Import Excel du personnel"})

            created = 0
            updated = 0
            errors = []

            with transaction.atomic():
                for line_number, row in enumerate(rows[1:], start=2):
                    if not any(row):
                        continue

                    def cell(name):
                        idx = get_index(name)
                        return row[idx] if idx is not None and idx < len(row) else None

                    first_name = _normalize_text(cell('prenom'))
                    last_name = _normalize_text(cell('nom'))
                    position_name = _normalize_text(cell('poste_fonction'))
                    category_name = _normalize_text(cell('categorie'))
                    education_level = _normalize_text(cell('niveau_d_etudes'))
                    matricule = _normalize_text(cell('matricule'))
                    photo_value = cell('photo')

                    if not all([first_name, last_name, position_name, category_name, education_level]):
                        errors.append(f"Ligne {line_number}: champs obligatoires manquants.")
                        continue

                    defaults = {
                        'first_name': first_name,
                        'last_name': last_name,
                        'position': Position.objects.get_or_create(name=position_name)[0],
                        'category': Category.objects.get_or_create(name=category_name)[0],
                        'education_level': education_level,
                    }
                    if matricule:
                        defaults['matricule'] = matricule

                    if photo_value:
                        # Le chargement de photo depuis Excel n'est pas pris en charge automatiquement.
                        pass

                    personnel = None
                    if matricule:
                        personnel = Personnel.objects.filter(matricule=matricule).first()
                    if personnel is None:
                        personnel = Personnel.objects.filter(
                            first_name__iexact=first_name,
                            last_name__iexact=last_name,
                        ).first()

                    if personnel:
                        for field, value in defaults.items():
                            setattr(personnel, field, value)
                        personnel.save()
                        updated += 1
                    else:
                        Personnel.objects.create(**defaults)
                        created += 1

            results = {'created': created, 'updated': updated, 'errors': errors}
            if created or updated:
                messages.success(
                    request,
                    f"Import terminé : {created} créé(s), {updated} mis à jour(s)."
                )
            if errors:
                messages.warning(request, f"{len(errors)} ligne(s) ont été ignorées.")

            return render(request, 'personnel/personnel_import.html', {
                'form': PersonnelImportForm(),
                'title': "Import Excel du personnel",
                'results': results,
            })
    else:
        form = PersonnelImportForm()

    return render(request, 'personnel/personnel_import.html', {'form': form, 'title': 'Import Excel du personnel'})

class PersonnelListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Personnel
    template_name = 'personnel/personnel_list.html'
    context_object_name = 'personnels'
    paginate_by = 10
    permission_required = 'cards.view_personnel'

class PersonnelDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = Personnel
    template_name = 'personnel/personnel_detail.html'
    context_object_name = 'personnel'
    permission_required = 'cards.view_personnel'

class PersonnelCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Personnel
    form_class = PersonnelForm
    template_name = 'personnel/personnel_form.html'
    success_url = reverse_lazy('cards:personnel_list')
    permission_required = 'cards.add_personnel'

    def get_initial(self):
        initial = super().get_initial()
        import datetime
        year = datetime.datetime.now().year
        last_personnel = Personnel.objects.order_by('-id').first()
        new_id = (last_personnel.id + 1) if last_personnel else 1
        initial['matricule'] = f"MAT-{year}-{new_id:04d}"
        return initial

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, "Personnel enregistré avec succès.")
        return response

class PersonnelUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Personnel
    form_class = PersonnelForm
    template_name = 'personnel/personnel_form.html'
    permission_required = 'cards.change_personnel'

    def get_success_url(self):
        return reverse('cards:personnel_detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, "Personnel mis à jour avec succès.")
        return response

class PersonnelDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = Personnel
    template_name = 'personnel/personnel_confirm_delete.html'
    success_url = reverse_lazy('cards:personnel_list')
    permission_required = 'cards.delete_personnel'

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, "Personnel supprimé avec succès.")
        return response

# --- CARD VIEWS ---

class CardListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Card
    template_name = 'cards/card_list.html'
    context_object_name = 'cards'
    paginate_by = 10
    permission_required = 'cards.view_card'

class CardDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = Card
    template_name = 'cards/card_detail.html'
    context_object_name = 'card'
    permission_required = 'cards.view_card'

class CardCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Card
    form_class = CardForm
    template_name = 'cards/card_form.html'
    success_url = reverse_lazy('cards:card_list')
    permission_required = 'cards.add_card'

    def form_valid(self, form):
        response = super().form_valid(form)
        try:
            generate_card_image(self.object.id, base_url=self.request.build_absolute_uri("/"))
            messages.success(self.request, "Carte créée et générée avec succès.")
        except Exception as e:
            messages.error(self.request, f"Carte créée mais erreur lors de la génération de l'image: {str(e)}")
        return response

class CardUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Card
    form_class = CardForm
    template_name = 'cards/card_form.html'
    permission_required = 'cards.change_card'
    
    def get_success_url(self):
        return reverse('cards:card_detail', kwargs={'pk': self.object.pk})
        
    def form_valid(self, form):
        response = super().form_valid(form)
        try:
            generate_card_image(self.object.id, base_url=self.request.build_absolute_uri("/"))
            messages.success(self.request, "Carte mise à jour et regénérée avec succès.")
        except Exception as e:
            messages.error(self.request, f"Mise à jour réussie mais erreur lors de la regénération: {str(e)}")
        return response

@login_required
@permission_required('cards.change_card', raise_exception=True)
def generate_card(request, pk):
    card = get_object_or_404(Card, pk=pk)
    try:
        generate_card_image(card.id, base_url=request.build_absolute_uri("/"))
        messages.success(request, "Image de la carte générée avec succès.")
    except Exception as e:
        messages.error(request, f"Erreur lors de la génération: {str(e)}")
    
    return redirect('cards:card_detail', pk=pk)


@login_required
@permission_required('cards.view_card', raise_exception=True)
def card_pdf(request, pk):
    card = get_object_or_404(
        Card.objects.select_related("personnel"),
        pk=pk,
    )

    if not card.generated_card:
        messages.error(request, "Veuillez d'abord générer la carte PVC avant de créer le PDF.")
        return redirect('cards:card_detail', pk=pk)

    try:
        from PIL import Image
    except ImportError as exc:
        messages.error(request, "La génération PDF nécessite Pillow.")
        return redirect('cards:card_detail', pk=pk)

    pdf_buffer = BytesIO()
    images = []

    recto_path = card.generated_card.path if hasattr(card.generated_card, "path") else None
    verso_path = card.generated_card_back.path if card.generated_card_back and hasattr(card.generated_card_back, "path") else None

    if recto_path:
        recto = Image.open(recto_path).convert("RGB")
        images.append(recto)

    if verso_path:
        verso = Image.open(verso_path).convert("RGB")
        images.append(verso)

    if not images:
        messages.error(request, "Aucune image de carte n’est disponible pour générer le PDF.")
        return redirect('cards:card_detail', pk=pk)

    first, *rest = images
    first.save(
        pdf_buffer,
        format="PDF",
        save_all=True,
        append_images=rest,
        resolution=300.0,
    )
    pdf_buffer.seek(0)

    filename = f"carte_{pk}.pdf"
    response = HttpResponse(pdf_buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def public_card_profile(request, public_token):
    try:
        card = Card.objects.select_related("personnel", "personnel__position", "personnel__category").get(
            public_token=public_token
        )
    except Card.DoesNotExist:
        return render(
            request,
            "cards/public_profile_unavailable.html",
            {"public_token": public_token, "today": timezone.localdate()},
            status=404,
        )

    if not card.personnel_id:
        return render(
            request,
            "cards/public_profile_unavailable.html",
            {"public_token": public_token, "today": timezone.localdate()},
            status=404,
        )

    personnel = card.personnel
    full_name = f"{personnel.first_name} {personnel.last_name}".strip()
    position_label = personnel.position.name if personnel.position else ""
    category_label = personnel.category.name if personnel.category else ""

    context = {
        "card": card,
        "personnel": personnel,
        "full_name": full_name,
        "position_label": position_label,
        "category_label": category_label,
        "today": timezone.localdate(),
    }
    return render(request, "cards/public_profile.html", context)

class CardDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = Card
    template_name = 'cards/card_confirm_delete.html'
    success_url = reverse_lazy('cards:card_list')
    permission_required = 'cards.delete_card'

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, "Carte supprimée avec succès.")
        return response
